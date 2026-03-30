"""Report Generation Engine - Excel and PDF reports."""

from __future__ import annotations

import io
import logging
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    Workbook = None
    logger.warning("openpyxl not installed")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        PageBreak,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    SimpleDocTemplate = None
    logger.warning("reportlab not installed")

MAX_ROWS_PER_SHEET = 50000
PREVIEW_ROWS = 100


class ExcelReportGenerator:
    """Excel report generator."""

    def __init__(self, data: list[dict[str, Any]] | pd.DataFrame):
        if isinstance(data, list):
            self.df = pd.DataFrame(data)
        else:
            self.df = data

    def generate(
        self,
        title: str = "Data Report",
        include_summary: bool = True,
        include_statistics: bool = True,
        max_rows: int = MAX_ROWS_PER_SHEET,
    ) -> bytes:
        """Generate Excel report."""
        if Workbook is None:
            raise ImportError("openpyxl not installed")

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells(f"A1:{chr(65 + min(len(self.df.columns), 10))}1")

        row = 3
        if include_summary:
            ws[f"A{row}"] = "Summary"
            ws[f"A{row}"].font = header_font
            row += 1

            ws[f"A{row}"] = "Total Rows"
            ws[f"B{row}"] = len(self.df)
            row += 1

            ws[f"A{row}"] = "Total Columns"
            ws[f"B{row}"] = len(self.df.columns)
            row += 1

            ws[f"A{row}"] = "Memory Usage"
            ws[f"B{row}"] = (
                f"{self.df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB"
            )
            row += 1

            ws[f"A{row}"] = "Completeness"
            ws[f"B{row}"] = f"{(1 - self.df.isna().mean().mean()) * 100:.1f}%"
            row += 2

        if include_statistics:
            ws[f"A{row}"] = "Column Statistics"
            ws[f"A{row}"].font = header_font
            row += 1

            stats_data = [["Column", "Type", "Non-Null", "Null %", "Unique"]]
            for col in self.df.columns[:20]:
                stats_data.append(
                    [
                        col,
                        str(self.df[col].dtype),
                        int(self.df[col].notna().sum()),
                        f"{self.df[col].isna().mean() * 100:.1f}%",
                        self.df[col].nunique(),
                    ]
                )

            for r_idx, row_data in enumerate(stats_data, start=row):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.font = header_font_white

            row += len(stats_data) + 2

        ws[f"A{row}"] = "Data Preview"
        ws[f"A{row}"].font = header_font
        row += 1

        preview_df = self.df.head(PREVIEW_ROWS)
        for c_idx, col in enumerate(preview_df.columns, start=1):
            cell = ws.cell(row=row, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.font = header_font_white

        for r_idx, row_data in enumerate(preview_df.values, start=row + 1):
            for c_idx, value in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=str(value)[:100])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class PDFReportGenerator:
    """PDF report generator."""

    def __init__(self, data: list[dict[str, Any]] | pd.DataFrame):
        if isinstance(data, list):
            self.df = pd.DataFrame(data)
        else:
            self.df = data

        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self) -> None:
        """Setup custom styles."""
        self.styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=self.styles["Heading1"],
                fontSize=18,
                alignment=TA_CENTER,
                spaceAfter=20,
            )
        )
        self.styles.add(
            ParagraphStyle(
                name="SectionHeader",
                parent=self.styles["Heading2"],
                fontSize=14,
                spaceAfter=10,
                spaceBefore=15,
            )
        )

    def generate(
        self,
        title: str = "Data Analytics Report",
        include_summary: bool = True,
        include_statistics: bool = True,
        include_recommendations: bool = True,
    ) -> bytes:
        """Generate PDF report."""
        if SimpleDocTemplate is None:
            raise ImportError("reportlab not installed")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter, topMargin=0.5 * inch, bottomMargin=0.5 * inch
        )

        story = []

        story.append(Paragraph(title, self.styles["CustomTitle"]))

        if include_summary:
            story.append(Paragraph("Executive Summary", self.styles["SectionHeader"]))
            summary_data = [
                ["Metric", "Value"],
                ["Total Records", str(len(self.df))],
                ["Total Columns", str(len(self.df.columns))],
                [
                    "Memory Usage",
                    f"{self.df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB",
                ],
                ["Completeness", f"{(1 - self.df.isna().mean().mean()) * 100:.1f}%"],
            ]
            summary_table = Table(summary_data, colWidths=[2.5 * inch, 2.5 * inch])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(summary_table)
            story.append(Spacer(1, 0.25 * inch))

        if include_statistics:
            story.append(Paragraph("Column Statistics", self.styles["SectionHeader"]))

            stats_data = [["Column", "Type", "Non-Null", "Unique"]]
            for col in self.df.columns[:15]:
                stats_data.append(
                    [
                        col[:20],
                        str(self.df[col].dtype)[:15],
                        str(int(self.df[col].notna().sum())),
                        str(self.df[col].nunique()),
                    ]
                )

            stats_table = Table(
                stats_data, colWidths=[2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch]
            )
            stats_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(stats_table)
            story.append(Spacer(1, 0.25 * inch))

        if include_recommendations:
            story.append(
                Paragraph(
                    "Key Findings & Recommendations", self.styles["SectionHeader"]
                )
            )

            findings = []

            null_ratio = self.df.isna().mean().mean()
            if null_ratio > 0.1:
                findings.append(
                    f"• Missing values detected: {null_ratio * 100:.1f}% of data"
                )

            dup_ratio = (
                self.df.duplicated().sum() / len(self.df) if len(self.df) > 0 else 0
            )
            if dup_ratio > 0.01:
                findings.append(f"• Duplicate rows found: {dup_ratio * 100:.1f}%")

            numeric_cols = self.df.select_dtypes(include=["number"]).columns
            if len(numeric_cols) > 0:
                findings.append(
                    f"• {len(numeric_cols)} numeric columns available for analysis"
                )

            if not findings:
                findings.append("• Data quality appears good")

            for finding in findings[:5]:
                story.append(Paragraph(finding, self.styles["BodyText"]))
                story.append(Spacer(1, 0.1 * inch))

        story.append(Spacer(1, 0.25 * inch))
        story.append(
            Paragraph(
                "<i>Generated by InsightFlow AI Analytics Platform</i>",
                ParagraphStyle(
                    "Footer",
                    parent=self.styles["Normal"],
                    fontSize=8,
                    textColor=colors.grey,
                ),
            )
        )

        doc.build(story)
        return buffer.getvalue()


def generate_excel_report(
    data: list[dict[str, Any]] | pd.DataFrame,
    title: str = "Data Report",
) -> bytes:
    """Convenience function to generate Excel report."""
    generator = ExcelReportGenerator(data)
    return generator.generate(title=title)


def generate_pdf_report(
    data: list[dict[str, Any]] | pd.DataFrame,
    title: str = "Data Analytics Report",
) -> bytes:
    """Convenience function to generate PDF report."""
    generator = PDFReportGenerator(data)
    return generator.generate(title=title)
